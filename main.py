import yfinance as yf
import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, timezone
from alpaca_trade_api.rest import REST
import requests
from bs4 import BeautifulSoup
from io import StringIO
import time

# Alpaca API konfigurációja
ALPACA_API_KEY = "PKEUE2RW2GRN346BC03T"
ALPACA_SECRET_KEY = "IUQsCnGkOBtrXKeoYl7Z2BDmSIWbEWtTkLI87NRP"
ALPACA_BASE_URL = "https://paper-api.alpaca.markets"

alpaca = REST(ALPACA_API_KEY, ALPACA_SECRET_KEY, base_url=ALPACA_BASE_URL)

# Hibák gyűjtése
errors = []


def fetch_sp500_tickers():
    try:
        url = "https://www.slickcharts.com/sp500"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        response = requests.get(url, headers=headers)
        df = pd.read_html(StringIO(response.text))[0]

        df['Weight'] = df['Weight'].str.rstrip('%').astype(float)
        df = df.sort_values('Weight', ascending=False)
        tickers = df['Symbol'].tolist()

        print(f"SlickCharts: {len(tickers)} ticker találva")
        return tickers

    except Exception as e:
        print(f"Hiba a SlickCharts adatok lekérésénél: {str(e)}")
        return []


def get_close_prices(data):
    if 'Close' in data.columns:
        return data['Close']
    elif 'close' in data.columns:
        return data['close']
    else:
        raise ValueError("Nem található záróár oszlop az adatokban")


def get_volume(data):
    if 'Volume' in data.columns:
        return data['Volume']
    elif 'volume' in data.columns:
        return data['volume']
    else:
        raise ValueError("Nem található volume oszlop az adatokban")


def calculate_exponential_regression(data, days=90):
    try:
        if len(data) < days:
            return 0, 0

        last_90_days = data[-days:].dropna()
        if len(last_90_days) < 2:
            return 0, 0

        x = np.arange(len(last_90_days))
        y = np.log(last_90_days.values)

        slope, intercept, r_value, _, _ = stats.linregress(x, y)
        annualized_slope = (np.exp(slope) ** 250) - 1
        return annualized_slope, r_value ** 2
    except Exception as e:
        print(f"Hiba a regresszió számításánál: {e}")
        return 0, 0


def fetch_stock_data(ticker, start_date, end_date):
    alpaca_ticker = ticker.replace('.', '-')

    try:
        time.sleep(0.1)  # Rate limiting elkerülése
        yesterday = end_date - timedelta(days=2)
        alpaca_data = alpaca.get_bars(
            alpaca_ticker,
            start=start_date.strftime('%Y-%m-%d'),
            end=yesterday.strftime('%Y-%m-%d'),
            timeframe='1Day',
            adjustment='raw',
            feed='iex'
        ).df

        if not alpaca_data.empty and len(alpaca_data) >= 90:
            # Volume és Close oszlopok egységesítése
            if 'volume' in alpaca_data.columns:
                alpaca_data.rename(columns={'volume': 'Volume', 'close': 'Close'}, inplace=True)
            print(f"Alpaca adatok használata: {ticker}")
            return alpaca_data
    except Exception as e:
        print(f"Alpaca hiba a {ticker} esetében: {str(e)}")

    # Yahoo Finance próbálkozás
    try:
        time.sleep(0.1)
        yahoo_ticker = ticker
        if '.B' in ticker:
            yahoo_ticker = ticker.replace('.B', '-B')

        stock = yf.Ticker(yahoo_ticker)
        data = stock.history(start=start_date, end=end_date, interval='1d')
        if not data.empty and len(data) >= 90:
            print(f"Yahoo Finance adatok használata: {ticker}")
            return data
    except Exception as e:
        print(f"Yahoo Finance hiba a {ticker} esetében: {str(e)}")

    errors.append([ticker, "Nem található adat egyik forrásból sem"])
    return None


def check_gap(data, threshold=0.15):
    try:
        if 'Open' in data.columns and 'Close' in data.columns:
            gaps = (data['Open'] - data['Close'].shift(1)) / data['Close'].shift(1)
        else:
            gaps = (data['open'] - data['close'].shift(1)) / data['close'].shift(1)
        return (gaps.abs() > threshold).any()
    except Exception as e:
        errors.append(["Gap Check", f"Hiba történt a gap ellenőrzése során: {e}"])
        return False


def create_excel_file(sp500_data, stock_data):
    try:
        wb = Workbook()
        filter_sheet = wb.active
        filter_sheet.title = "Filter"
        database_sheet = wb.create_sheet("Database")
        error_sheet = wb.create_sheet("Error")

        filter_headers = ["S&P 500 > 200 MA", "Ticker", "Regression * R-squared",
                        "Market Cap", "Above 100 MA", "15% Gap"]
        database_headers = ["Ticker", "90-day Annualized Exp Regression",
                            "R-squared", "Regression * R-squared", "Close"]

        for col, header in enumerate(filter_headers, start=1):
            filter_sheet.cell(row=1, column=col, value=header)
        for col, header in enumerate(database_headers, start=1):
            database_sheet.cell(row=1, column=col, value=header)

        ma_condition = sp500_data['Close'].iloc[-1] > sp500_data['Close'].rolling(window=200).mean().iloc[-1]
        filter_sheet['A2'] = "YES" if ma_condition.item() else "NO"
        filter_sheet['A2'].fill = PatternFill(
            start_color="00FF00" if ma_condition.item() else "FF0000",
            fill_type="solid"
        )

        sorted_stocks = sorted(
            stock_data.items(),
            key=lambda x: x[1]['reg'] * x[1]['r_squared'],
            reverse=True
        )

        for row, (ticker, data) in enumerate(sorted_stocks, start=2):
            database_sheet.cell(row=row, column=1, value=ticker)
            database_sheet.cell(row=row, column=2, value=data['reg'])
            database_sheet.cell(row=row, column=3, value=data['r_squared'])
            database_sheet.cell(row=row, column=4, value=data['reg'] * data['r_squared'])
            database_sheet.cell(row=row, column=5, value=data['close'])

            filter_sheet.cell(row=row, column=2, value=ticker)
            filter_sheet.cell(row=row, column=3, value=data['reg'] * data['r_squared'])
            filter_sheet.cell(row=row, column=4, value=data['market_cap'])

            # Above 100 MA oszlop színezése (E oszlop) - YES=zöld, NO=piros
            cell_e = filter_sheet.cell(row=row, column=5, value="YES" if data['above_ma100'] else "NO")
            cell_e.fill = PatternFill(
                start_color="00FF00" if data['above_ma100'] else "FF0000",
                fill_type="solid"
            )

            # 15% Gap oszlop színezése (F oszlop) - YES=piros, NO=zöld
            cell_f = filter_sheet.cell(row=row, column=6, value="YES" if data['gap'] else "NO")
            cell_f.fill = PatternFill(
                start_color="FF0000" if data['gap'] else "00FF00",
                fill_type="solid"
            )


        error_sheet.append(["Ticker", "Hibaüzenet"])
        for error in errors:
            error_sheet.append(error)

        wb.save("momentum_strategy_output.xlsx")
        print("Excel fájl sikeresen létrehozva.")
    except Exception as e:
        print(f"Hiba történt az Excel fájl létrehozása során: {e}")


def main():
    try:
        print("S&P 500 adatok letöltése")
        sp500_data = yf.download("^GSPC", period="1y", interval="1d")

        print("S&P 500 komponensek lekérése")
        sp500_tickers = fetch_sp500_tickers()

        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=180)

        stock_data = {}
        total_tickers = len(sp500_tickers)
        processed_tickers = 0

        print(f"Összesen {total_tickers} részvény feldolgozása")
        for ticker in sp500_tickers:
            try:
                data = fetch_stock_data(ticker, start_date, end_date)
                if data is not None and not data.empty:
                    close_prices = get_close_prices(data)
                    volume = get_volume(data)
                    reg, r_squared = calculate_exponential_regression(close_prices)
                    stock_data[ticker] = {
                        'reg': reg,
                        'r_squared': r_squared,
                        'market_cap': str(round((yf.Ticker(ticker).info['marketCap'] / 1000000000), 2)) + " blns$",
                        'above_ma100': close_prices.iloc[-1] > close_prices.rolling(window=100).mean().iloc[-1],
                        'gap': check_gap(data),
                        'close': close_prices.iloc[-1]
                    }
                    processed_tickers += 1
                    progress = (processed_tickers / total_tickers) * 100
                    print(f"Feldolgozás: {progress:.2f}% ({processed_tickers}/{total_tickers}) - {ticker}")
            except Exception as e:
                errors.append([ticker, f"Hiba történt a feldolgozás során: {str(e)}"])

        create_excel_file(sp500_data, stock_data)
    except Exception as e:
        print(f"Általános hiba történt: {str(e)}")


if __name__ == "__main__":
    main()
